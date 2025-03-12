import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class DataExporter:
    def __init__(self, config):
        self.config = config
    
    def export_to_excel(self, video_info):
        """导出视频信息到Excel"""
        excel_path = self.config.get('excel_path')
        
        # 检查Excel文件路径
        if not excel_path:
            excel_path = os.path.join(self.config.get('download_path'), 'video_info.xlsx')
            self.config.set('excel_path', excel_path)
        
        # 准备导出数据
        data = {
            '视频标题': video_info.get('title', ''),
            '视频描述': video_info.get('description', ''),
            '标签': '、'.join(video_info.get('tags', [])),
            '文字稿': video_info.get('transcript', ''),
            '点赞数': video_info.get('stats', {}).get('likes', 0),
            '评论数': video_info.get('stats', {}).get('comments', 0),
            '收藏数': video_info.get('stats', {}).get('favorites', 0),
            '转发数': video_info.get('stats', {}).get('shares', 0),
            '账号名称': video_info.get('author', {}).get('name', ''),
            '账号ID': video_info.get('author', {}).get('id', ''),
            '源链接': video_info.get('source_url', '')
        }
        
        # 检查Excel文件是否已存在
        if os.path.exists(excel_path):
            try:
                # 加载现有Excel文件
                df_existing = pd.read_excel(excel_path)
                
                # 添加新行
                df_new = pd.DataFrame([data])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                
                # 保存到Excel
                df_combined.to_excel(excel_path, index=False)
                
                # 调整列宽
                self._adjust_column_width(excel_path)
                
                return {'success': True, 'message': f'数据已导出到 {excel_path}'}
            except Exception as e:
                return {'success': False, 'message': f'导出数据失败: {str(e)}'}
        else:
            try:
                # 创建新的Excel文件
                df = pd.DataFrame([data])
                
                # 确保目录存在
                os.makedirs(os.path.dirname(excel_path), exist_ok=True)
                
                # 保存到Excel
                df.to_excel(excel_path, index=False)
                
                # 调整列宽
                self._adjust_column_width(excel_path)
                
                return {'success': True, 'message': f'数据已导出到 {excel_path}'}
            except Exception as e:
                return {'success': False, 'message': f'导出数据失败: {str(e)}'}
    
    def _adjust_column_width(self, excel_path):
        """调整Excel列宽"""
        try:
            # 加载工作簿
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # 设置列宽
            column_widths = {
                'A': 30,  # 视频标题
                'B': 50,  # 视频描述
                'C': 20,  # 标签
                'D': 50,  # 文字稿
                'E': 10,  # 点赞数
                'F': 10,  # 评论数
                'G': 10,  # 收藏数
                'H': 10,  # 转发数
                'I': 20,  # 账号名称
                'J': 20,  # 账号ID
                'K': 30,  # 源链接
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置自动换行
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrapText=True, vertical='top')
            
            # 保存工作簿
            wb.save(excel_path)
        except Exception as e:
            print(f"调整Excel列宽失败: {e}")